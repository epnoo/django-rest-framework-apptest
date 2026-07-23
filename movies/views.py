import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Director, Genre, Movie
from .serializers import DirectorSerializer, GenreSerializer, MovieSerializer

class DirectorViewSet(viewsets.ModelViewSet):
    queryset = Director.objects.all()
    serializer_class = DirectorSerializer

class GenreViewSet(viewsets.ModelViewSet):
    queryset = Genre.objects.all()
    serializer_class = GenreSerializer

class MovieViewSet(viewsets.ModelViewSet):
    queryset = Movie.objects.all()
    serializer_class = MovieSerializer

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        # Create a new Excel workbook and sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Movies Data'

        # Write header row
        headers = ['ID', 'Title', 'Release Year', 'Rating', 'Director ID', 'Genre ID']
        worksheet.append(headers)

        # Write data rows
        movies_query = Movie.objects.all()

        for movie in movies_query:
            worksheet.append([movie.id, 
                              movie.title, 
                              movie.release_year, 
                              movie.rating, 
                              movie.director.id, 
                              movie.genres.id])

        # Prepare the response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=movies.xlsx'
        workbook.save(response)
        return response

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active

        imported_count = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skip
            if row[1]:
                Movie.objects.create(
                    title=row[1],
                    release_year=row[2],
                    rating=row[3],
                    director_id=row[4],
                    genres_id=row[5]
                )
                imported_count += 1

        return Response({'message': f'{imported_count} movies imported successfully'}, status=status.HTTP_201_CREATED)